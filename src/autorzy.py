""" autorzy.xlsx -> QuickStatements """

import sys
from pathlib import Path
from openpyxl import load_workbook


P_INSTANCE_OF = 'P47'
Q_HUMAN = 'Q32'
P_IMIE = 'P3'
P_NAZWISKO = 'P4'

class Autor:
    """" klasa Autor """

    def __init__(self, p_etykieta: str = '', alias: str = '', p_imie: str = '',
                 p_nazwisko: str = ''):
        """ init """
        self.etykieta = p_etykieta.strip()
        self._alias = alias
        self.imie = p_imie.strip()
        self. nazwisko = p_nazwisko.strip()

    @property
    def alias(self):
        """ alias """
        return self._alias

    @alias.setter
    def alias(self, value: str):
        """ alias """
        if value:
            value = ' '.join(value.strip().split())
            lista = value.split("|")
            lista = [change_name_forname(item) for item in lista]
            self._alias = lista
        else:
            self._alias = []


def change_name_forname(name: str) -> str:
    """ change_name_forname"""
    name_parts = name.split(" ")
    if len(name_parts) == 2 and name_parts[0][0].isupper() and name_parts[1][0].isupper():
        result_name = name_parts[1] + " " + name_parts[0]
    elif (len(name_parts) == 3 and name_parts[0][0].isupper() and name_parts[1][0].isupper()
              and name_parts[2][0].isupper()):
        result_name = name_parts[1] + " " + name_parts[2] + " " + name_parts[0]
    else:
        print(f"ERROR: {name}")
    return result_name


if __name__ == "__main__":
    xlsx_path = Path('.').parent / 'data/autorzy.xlsx'
    output = Path('.').parent / 'out/autorzy.qs'

    try:
        wb = load_workbook(xlsx_path)
    except IOError:
        print(f"ERROR. Can't open and process file: {xlsx_path}")
        sys.exit(1)

    ws = wb['Arkusz1']

    col_names = {'NAZWA WŁAŚCIWA':0, 'NAZWA WARIANTYWNA (znany też jako)':1}

    autor_list = []
    for row in ws.iter_rows(2, ws.max_row):
        osoba = row[col_names['NAZWA WŁAŚCIWA']].value
        osoba_alias = row[col_names['NAZWA WARIANTYWNA (znany też jako)']].value

        if osoba:
            autor = Autor()
            osoba = ' '.join(osoba.strip().split()) # podwójne, wiodące i kończące spacje
            tmp = osoba.split(" ")

            if len(tmp) == 2 and tmp[0][0].isupper() and tmp[1][0].isupper():
                autor.etykieta = tmp[1] + " " + tmp[0]
                autor.imie = tmp[1]
                # jeżeli znamy tylko inicjał imienia to nie zakładamy Q
                if len(autor.imie) == 2 and autor.imie.endswith("."):
                    continue
                autor.nazwisko = tmp[0]
            elif (len(tmp) == 3 and tmp[0][0].isupper() and tmp[1][0].isupper()
                    and tmp[2][0].isupper()):
                autor.etykieta = tmp[1] + " " + tmp[2] + " " + tmp[0]
                autor.imie = tmp[1]
                autor.nazwisko = tmp[0]
            elif tmp[0].startswith('d’'):
                autor.etykieta = tmp[1] + " " + tmp[0]
                autor.imie = tmp[1]
                autor.nazwisko = tmp[0]
            elif "Szturm de Sztrem" in osoba:
                autor.etykieta = "Tadeusz Szturm de Sztrem"
                autor.imie = "Tadeusz"
                autor.nazwisko = "Szturm de Sztrem"
            else:
                print(f'ERROR: {osoba}')

            if osoba_alias:
                autor.alias = osoba_alias

            autor_list.append(autor)

    with open(output, "w", encoding='utf-8') as f:
        for autor in autor_list:
            f.write('CREATE\n')
            f.write(f'LAST\tLpl\t"{autor.etykieta}"\n')
            f.write(f'LAST\tLen\t"{autor.etykieta}"\n')
            f.write(f'LAST\t{P_INSTANCE_OF}\t{Q_HUMAN}\n')
            f.write(f'LAST\t{P_IMIE}\t"{autor.imie}"\n')
            f.write(f'LAST\t{P_NAZWISKO}\t"{autor.nazwisko}"\n')
            if autor.alias:
                for item in autor.alias:
                    f.write(f'LAST\tApl\t"{item}"\n')
                    f.write(f'LAST\tAen\t"{item}"\n')