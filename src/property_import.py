""" skrypt importujący właściwości z xlsx do wikibase """

import os
import sys
import re
from pathlib import Path
from typing import Union
from openpyxl import load_workbook
from wikibaseintegrator import wbi_core
from wikibaseintegrator.wbi_config import config as wbi_config
from wikibaseintegrator import wbi_login, wbi_datatype
from wikibaseintegrator.wbi_functions import mediawiki_api_call_helper
from wikibaseintegrator.wbi_exceptions import (MWApiError)
from dotenv import load_dotenv
from wikidariahtools import element_search


# adresy dla API Wikibase
wbi_config['MEDIAWIKI_API_URL'] = 'https://prunus-208.man.poznan.pl/api.php'
wbi_config['SPARQL_ENDPOINT_URL'] = 'https://prunus-208.man.poznan.pl/bigdata/sparql'
wbi_config['WIKIBASE_URL'] = 'https://prunus-208.man.poznan.pl'

# słownik globalnych referencji dla arkuszy (z deklaracjami)
GLOBAL_REFERENCE = {}


# --- klasy ---
class BasicProp:
    """ Identyfikatory podstawowych właściwości
    """
    def __init__(self):
        self.wiki_id = ''
        self.wiki_url = ''
        self.inverse = ''

    def get_wiki_properties(self):
        """ funkcja ustala nr podstawowych property związanych z wikidata.org
        """
        if self.wiki_id == '':
            search_result, pid = element_search('Wikidata ID', 'property', 'en')
            if search_result:
                self.wiki_id = pid
        if self.wiki_url == '':
            search_result, pid = element_search('Wikidata URL', 'property', 'en')
            if search_result:
                self.wiki_url = pid
        if self.inverse == '':
            search_result, pid = element_search('inverse property', 'property', 'en')
            if search_result:
                self.inverse = pid


class WDHSpreadsheet:
    """ Plik arkusza kalkulacyjnego z modelem danych dla Wikibase
    """
    def __init__(self, path: str):
        self.path = path
        self.sheets = ['P_list', 'P_statements', 'Q_list', 'Q_statements', 'Globals']
        self.p_list = None          # arkusz z listą właściwości
        self.p_statements = None    # arkusz z listą deklaracji dla właściwości
        self.workbook = None
        self.property_columns = []
        self.statement_columns = []
        self.i_list = None          # arkusz z listą elementów (item)
        self.i_statements = None    # arkusz z listą deklaracji dla elementów
        self.item_columns = []
        self.item_statement_columns = []
        self.globals = None         # arkusz z globalnymi referencjami
        self.globals_columns = []

    @property
    def path(self) -> str:
        """ get path """
        return self._path

    @path.setter
    def path(self, value: str):
        """ set path """
        self._path = value

    def open(self):
        """ odczyt pliku i weryfikacja poprawności """
        try:
            self.workbook = load_workbook(self.path)
        except IOError:
            print(f"ERROR. Can't open and process file: {self.path}")
            sys.exit(1)

        # czy to jest właściwy plik? cz. 1
        for sheet in self.sheets:
            if not sheet in self.workbook.sheetnames:
                print(f"ERROR. Expected worksheet '{sheet}' is missing in the file.")
                sys.exit(1)

        # arkusz właściwości
        self.p_list = self.workbook[self.sheets[0]]
        self.property_columns = self.get_col_names(self.p_list)
        p_list_expected = ['Label_en', 'Description_en', 'datatype', 'Label_pl']
        res, inf = self.test_columns(self.property_columns, p_list_expected)
        if not res:
            print(f'ERROR. Worksheet {self.sheets[0]}. The expected columns ({inf}) are missing.')
            sys.exit(1)

        # arkusz deklaracji dla właściwości
        self.p_statements = self.workbook[self.sheets[1]]
        self.statement_columns = self.get_col_names(self.p_statements)
        p_statements_expected = ['Label_en', 'P', 'value', 'reference_property', 'reference_value']
        res, inf = self.test_columns(self.statement_columns, p_statements_expected)
        if not res:
            print(f'ERROR. Worksheet {self.sheets[1]}. The expected columns ({inf}) are missing.')
            sys.exit(1)

        # arkusz elementów
        self.i_list = self.workbook[self.sheets[2]]
        self.item_columns = self.get_col_names(self.i_list)
        i_list_expected = ['Label_en', 'Label_pl', 'Description_en', 'Description_pl', 'Wiki_id']
        res, inf = self.test_columns(self.item_columns, i_list_expected)
        if not res:
            print(f'ERROR. Worksheet {self.sheets[2]}. The expected columns ({inf}) are missing.')
            sys.exit(1)

        # arkusz deklaracji dla elementów
        self.i_statements = self.workbook[self.sheets[3]]
        self.item_statement_columns = self.get_col_names(self.i_statements)
        i_statements_expected = ['Label_en', 'P', 'Value','Qualifier', 'Qualifier_value']
        res, inf = self.test_columns(self.item_statement_columns, i_statements_expected)
        if not res:
            print(f'ERROR. Worksheet {self.sheets[3]}. The expected columns ({inf}) are missing.')
            sys.exit(1)

        # arkusz globalnych referencji dla poszczególnych arkuszy
        self.globals = self.workbook[self.sheets[4]]
        self.globals_columns = self.get_col_names(self.globals)
        globals_expected = ['Sheet', 'reference_property', 'reference_value']
        res, inf = self.test_columns(self.globals_columns, globals_expected)
        if not res:
            print(f'ERROR. Worksheet {self.sheets[4]}. The expected columns ({inf}) are missing.')
            sys.exit(1)


    def test_columns(self, t_col_names: dict, expected: list) -> tuple:
        """ weryfikuje czy arkusz zawiera oczekiwane kolumny """
        missing_cols = []
        res = True
        for col in expected:
            if not col in t_col_names:
                missing_cols.append(col)
                res = False

        return res, ",".join(missing_cols)


    def get_col_names(self, sheet) -> dict:
        """ funkcja zwraca słownik nazw kolumn
        """
        names = {}
        nr_col = 0
        for column in sheet.iter_cols(1, sheet.max_column):
            names[column[0].value] = nr_col
            nr_col += 1

        return names


    def correct_type(self, t_datatype: str) -> str:
        """ Funkcja ewentualnie koryguje typ właściwości na właściwy, zgodny z oczekiwanym
            przez Wikibase
        """
        if t_datatype is not None:
            if t_datatype == 'item':
                t_datatype = 'wikibase-item'
            elif t_datatype == 'property':
                t_datatype = 'wikibase-property'
            elif t_datatype == 'external identifier':
                t_datatype = 'external-id'
            elif value == 'URL':
                value = 'url'
            elif value == 'monolingual text':
                value = 'monolingualtext'
            elif value == 'geographic coordinates':
                value = 'globe-coordinate'
            elif value == 'point in time':
                value = 'time'

        return t_datatype

    def get_property_list(self) -> list:
        """ zwraca listę właściwości (w formie obiektów WDHProperty) do dodania
        """
        p_list = []
        for row in self.p_list.iter_rows(2, self.p_list.max_row):
            basic_cols = ['Label_en', 'Description_en', 'datatype', 'Label_pl']
            p_item = {}
            p_item = WDHProperty()
            for col in basic_cols:
                key = col.lower()
                col_value = row[self.property_columns[col]].value
                if key == 'label_en':
                    p_item.label_en = col_value
                elif key == 'description_en':
                    p_item.description_en = col_value
                elif key == 'datatype':
                    p_item.datatype = col_value
                elif key == 'label_pl':
                    p_item.label_pl = col_value

            # tylko jeżeli etykieta i opis w języku angielskim oraz typ danych są wypełnione
            # dane właściwości są dodawane do listy
            if p_item.label_en and p_item.description_en and p_item.datatype and p_item.label_pl:
                extend_cols = ['Description_pl', 'Wiki_id', 'inverse_property']
                for col in extend_cols:
                    key = col.lower()
                    col_value = row[self.property_columns[col]].value
                    if key == 'description_pl':
                        p_item.description_pl = col_value
                    elif key == 'wiki_id':
                        p_item.wiki_id = col_value
                    elif key == 'inverse_property':
                        p_item.inverse_property = col_value

                p_list.append(p_item)

        return p_list

    def get_statement_list(self) -> list:
        """ zwraca listę obiektów deklaracji do dodania
        """
        s_list = []
        for row in self.p_statements.iter_rows(2, self.p_statements.max_row):
            basic_cols = ['Label_en', 'P', 'value', 'reference_property', 'reference_value']
            s_item = WDHStatementProperty()
            reference_property = reference_value = ''
            for col in basic_cols:
                key = col.lower()
                col_value = row[self.statement_columns[col]].value

                if key == 'label_en':
                    s_item.label_en = col_value
                elif key == 'p':
                    s_item.statement_property = col_value
                elif key == 'value':
                    if not isinstance(col_value, str):
                        col_value = str(col_value)
                    s_item.statement_value = col_value                    
                elif key == 'reference_property':
                    reference_property = col_value
                elif key == 'reference_value':
                    reference_value = col_value

                if reference_property and reference_value:
                    s_item.references[reference_property] = reference_value

            # nazwa arkusza z deklaracjami dla właściwości
            s_item.sheet_name = self.sheets[1]

            # jeżeli są globalne referencje
            if s_item.sheet_name in GLOBAL_REFERENCE:
                g_ref_property, g_ref_value = GLOBAL_REFERENCE[s_item.sheet_name]
                s_item.additional_references[g_ref_property] = g_ref_value

            # tylko jeżeli etykieta w języku angielskim, właściwość i wartość są wypełnione
            # dane deklaracji są dodawane do listy
            if s_item.label_en and s_item.statement_property and s_item.statement_value:
                s_list.append(s_item)
            # jeżeli nie ma wartości etykiety, właściwości i wartości deklaracji
            # a są dane referencji to  dodaje referencje do ostatnio dodanej
            # pozycji z listy
            else:
                if reference_property and reference_value:
                    s_list[-1].references[reference_property] = reference_value

        return s_list

    def get_item_list(self) -> list:
        """ zwraca listę elementów (w formie obiektów WDHItem) do dodania
        """
        i_list = []
        for row in self.i_list.iter_rows(2, self.i_list.max_row):
            basic_cols = ['Label_en', 'Label_pl', 'Description_en', 'Description_pl']
            i_item = WDHItem()
            for col in basic_cols:
                key = col.lower()
                col_value = row[self.item_columns[col]].value
                if key == 'label_en':
                    i_item.label_en = col_value
                elif key == 'description_en':
                    i_item.description_en = col_value
                elif key == 'description_pl':
                    i_item.description_pl = col_value
                elif key == 'label_pl':
                    i_item.label_pl = col_value

            # tylko jeżeli etykieta i opis w języku angielskim oraz polskim są wypełnione
            # dane właściwości są dodawane do listy
            if i_item.label_en and i_item.description_en and i_item.description_pl and i_item.label_pl:
                extend_cols = ['Wiki_id']
                for col in extend_cols:
                    key = col.lower()
                    col_value = row[self.item_columns[col]].value
                    if key == 'wiki_id':
                        i_item.wiki_id = col_value

                i_list.append(i_item)

        return i_list

    def get_item_statement_list(self) -> list:
        """ zwraca listę obiektów deklaracji do dodania do elementów
        """

        s_list = []
        for row in self.i_statements.iter_rows(2, self.i_statements.max_row):
            basic_cols = ['Label_en', 'P', 'Value', 'Qualifier', 'Qualifier_value']

            label_en = statement_property = statement_value = qualifier = qualifier_value = ''
            for col in basic_cols:
                key = col.lower()
                col_value = row[self.item_statement_columns[col]].value

                if key == 'label_en':
                    label_en = col_value
                elif key == 'p':
                    statement_property = col_value
                elif key == 'value':
                    statement_value = col_value
                elif key == 'qualifier':
                    qualifier = col_value
                elif key == 'qualifier_value':
                    qualifier_value = col_value

            # tylko jeżeli etykieta w języku angielskim, właściwość i wartość są wypełnione
            # dane deklaracji są dodawane do listy
            if label_en and statement_property and statement_value:
                s_item = WDHStatementItem()
                s_item.label_en = label_en
                s_item.statement_property = statement_property
                s_item.statement_value = statement_value
                if qualifier and qualifier_value:
                    s_item.qualifiers[qualifier] = qualifier_value
                s_item.sheet_name = self.sheets[3]
                
                # jeżeli są globalne referencje
                if s_item.sheet_name in GLOBAL_REFERENCE:
                    g_ref_property, g_ref_value = GLOBAL_REFERENCE[s_item.sheet_name]
                    s_item.additional_references[g_ref_property] = g_ref_value

                s_list.append(s_item)
            # jeżeli nie ma wartości etykiety, właściwości i wartości deklaracji
            # a są dane kwalifikatora to  dodaje kwalifikator do ostatnio dodanej
            # pozycji z listy
            else:
                if qualifier and qualifier_value:
                    s_list[-1].qualifiers[qualifier] = qualifier_value

        return s_list

    def get_global(self) -> dict:
        """ get_global """
        global GLOBAL_REFERENCE

        for row in self.globals.iter_rows(2, self.globals.max_row):
            g_sheet = row[self.globals_columns['Sheet']].value
            g_property = row[self.globals_columns['reference_property']].value
            g_value = row[self.globals_columns['reference_value']].value
            GLOBAL_REFERENCE[g_sheet] = (g_property, g_value)



class WDHProperty:
    """ Klasa dla właściwości (property)
    """
    def __init__(self, label_en: str = '', description_en: str = '', datatype: str = '',
                 label_pl: str = '', description_pl: str = '', wiki_id: str = '',
                 inverse_property: str = ''):
        self.label_en = label_en
        self.description_en = description_en
        self.datatype = datatype
        self.label_pl = label_pl
        self.description_pl = description_pl
        self.wiki_id = wiki_id
        self.inverse_property = inverse_property

    @property
    def label_en(self) -> str:
        """ get label_en """
        return self._label_en

    @label_en.setter
    def label_en(self, value: str):
        """ set label_en """
        if value:
            self._label_en = value.strip()
        else:
            self._label_en = ''

    @property
    def description_en(self) -> str:
        """ get description_en """
        return self._description_en

    @description_en.setter
    def description_en(self, value: str):
        """ set description_en """
        if value:
            self._description_en = value.strip()
        else:
            self._description_en = ''

    @property
    def datatype(self) -> str:
        """ get datatype """
        return self._datatype

    @datatype.setter
    def datatype(self, value: str):
        """ set datatype """
        if value:
            if value == 'item':
                value = 'wikibase-item'
            elif value == 'property':
                value = 'wikibase-property'
            elif value == 'external identifier':
                value = 'external-id'
            elif value == 'URL':
                value = 'url'
            elif value == 'monolingual text':
                value = 'monolingualtext'
            elif value == 'geographic coordinates':
                value = 'globe-coordinate'
            elif value == 'point in time':
                value = 'time'

            self._datatype = value.strip()
        else:
            self._datatype = ''

    @property
    def label_pl(self) -> str:
        """ get label_pl """
        return self._label_pl

    @label_pl.setter
    def label_pl(self, value: str):
        """ set label_pl """
        if value:
            self._label_pl = value.strip()
        else:
            self._label_pl = ''

    @property
    def description_pl(self) -> str:
        """ get description_pl """
        return self._description_pl

    @description_pl.setter
    def description_pl(self, value: str):
        """ set description_pl """
        if value:
            self._description_pl = value.strip()
        else:
            self._description_pl = ''

    @property
    def wiki_id(self) -> str:
        """ get wiki_id """
        return self._wiki_id

    @wiki_id.setter
    def wiki_id(self, value: str):
        """ set wiki_id """
        if value:
            self._wiki_id = value.strip()
        else:
            self._wiki_id = ''

    @property
    def inverse_property(self) -> str:
        """ get inverse_property """
        return self._inverse_property

    @inverse_property.setter
    def inverse_property(self, value: str):
        """ set inverse_property """
        if value:
            self._inverse_property = value.strip()
        else:
            self._inverse_property = ''

    def write_to_wikibase(self):
        """ zapis właściwości w instancji wikibase """
        pass


class WDHStatementProperty:
    """ Klasa dla deklaracji (statement) dla właściwości
    """
    def __init__(self, label_en: str = '', statement_property: str = '',
                 statement_value: str = '', reference_property: str = '',
                 reference_value: str = ''):
        self.label_en = label_en
        self.statement_property = statement_property
        self.statement_value = statement_value
        #self.reference_property = reference_property
        #self.reference_value = reference_value
        self.references = {}
        if reference_property and reference_value:
            self.references[reference_property.strip()] = reference_value.strip()
        self.sheet_name = ''
        self.additional_references = {}

    @property
    def label_en(self) -> str:
        """ getter: label_en """
        return self._label_en

    @label_en.setter
    def label_en(self, value: str):
        """ setter: label_en """
        if value:
            self._label_en = value.strip()
        else:
            self._label_en = ''

    @property
    def statement_property(self) -> str:
        """ get statement_property """
        return self._statement_property

    @statement_property.setter
    def statement_property(self, value: str):
        """ set statement_property"""
        if value:
            self._statement_property = value.strip()
        else:
            self._statement_property = ''

    @property
    def statement_value(self):
        """ get statement_value """
        return self._statement_value

    @statement_value.setter
    def statement_value(self, value: str):
        """ set statement_value"""
        if value:
            self._statement_value = value.strip()
        else:
            self._statement_value = ''

    @property
    def reference_property(self):
        """ get reference_property """
        return self._reference_property

    @reference_property.setter
    def reference_property(self, value: str):
        """ set reference_property"""
        if value:
            self._reference_property = value.strip()
        else:
            self._reference_property = ''

    @property
    def reference_value(self):
        """ gettet: reference_value """
        return self._reference_value

    @reference_value.setter
    def reference_value(self, value: str):
        """ setter: reference_value"""
        if value:
            self._reference_value = value.strip()
        else:
            self._reference_property = ''

    def write_to_wikibase(self):
        """ zapis deklaracji w instancji wikibase """
        pass



class WDHItem:
    """ Klasa dla elementu (item)
    """
    def __init__(self, label_en: str = '', description_en: str = '',
                 label_pl: str = '', description_pl: str = '', wiki_id: str = ''):
        self.label_en = label_en
        self.description_en = description_en
        self.label_pl = label_pl
        self.description_pl = description_pl
        self.wiki_id = wiki_id

    @property
    def label_en(self) -> str:
        """ get label_en """
        return self._label_en

    @label_en.setter
    def label_en(self, value: str):
        """ set label_en """
        if value:
            self._label_en = value.strip()
        else:
            self._label_en = ''

    @property
    def description_en(self) -> str:
        """ get description_en """
        return self._description_en

    @description_en.setter
    def description_en(self, value: str):
        """ set description_en """
        if value:
            self._description_en = value.strip()
        else:
            self._description_en = ''

    @property
    def label_pl(self) -> str:
        """ get label_pl """
        return self._label_pl

    @label_pl.setter
    def label_pl(self, value: str):
        """ set label_pl """
        if value:
            self._label_pl = value.strip()
        else:
            self._label_pl = ''

    @property
    def description_pl(self) -> str:
        """ get description_pl """
        return self._description_pl

    @description_pl.setter
    def description_pl(self, value: str):
        """ set description_pl """
        if value:
            self._description_pl = value.strip()
        else:
            self._description_pl = ''

    @property
    def wiki_id(self) -> str:
        """ get wiki_id """
        return self._wiki_id

    @wiki_id.setter
    def wiki_id(self, value: str):
        """ set wiki_id """
        if value:
            self._wiki_id = value.strip()
        else:
            self._wiki_id = ''

    def write_to_wikibase(self):
        """ zapis elementu w instancji wikibase """
        search_item, search_id = element_search(self.label_en, 'item', 'en', 
                                                description=self.description_en)
        if search_item:
            print(f"Item: '{self.label_en}' already exists: {search_id}, update mode enabled.")
            wd_item = wbi_core.ItemEngine(item_id=search_id)
            mode = 'updated: '
        else:
            wd_item = wbi_core.ItemEngine(new_item=True)
            mode = 'added: '

        wd_item.set_label(self.label_en, lang='en')
        wd_item.set_description(self.description_en, lang='en')
        wd_item.set_label(self.label_pl,lang='pl')
        wd_item.set_description(self.description_pl, lang='pl')
        wiki_dane = None
        if self.wiki_id:
            if (wikibase_prop.wiki_id == '' or wikibase_prop.wiki_url == ''):
                wikibase_prop.get_wiki_properties()
            url = f'https://www.wikidata.org/wiki/{self.wiki_id}'
            references = [
                [
                wbi_datatype.Url(value=url, prop_nr=wikibase_prop.wiki_url, is_reference=True)
                ]
            ]
            wiki_dane = wbi_datatype.ExternalID(value=self.wiki_id, prop_nr=wikibase_prop.wiki_id,
                references=references)

        # zapis w Wikibase
        try:
            new_id = wd_item.write(login_instance, entity_type='item')
            if search_item:
                new_id = search_id

            # deklaracje dla elementu
            data = []
            if wiki_dane:
                data.append(wiki_dane)
                wd_statement = wbi_core.ItemEngine(item_id=new_id, data=data, debug=False)
                wd_statement.write(login_instance, entity_type='item')

            print(mode + new_id)
        except (MWApiError, KeyError):
            print('ERROR: ', self.label_en)


class WDHStatementItem:
    """ Klasa dla deklaracji (statement) dla elementów
    """
    def __init__(self, label_en: str = '', statement_property: str = '',
                 statement_value: str = '', qualifier: str = '',
                 qualifier_value: str = ''):
        self.label_en = label_en
        self.statement_property = statement_property
        self.statement_value = statement_value
        self.qualifiers = {}
        if qualifier and qualifier_value:
            self.qualifiers[qualifier.strip()] = qualifier_value.strip()
        self.sheet_name = ''
        self.references = {}
        self.additional_references = {}

    @property
    def label_en(self) -> str:
        """ getter: label_en """
        return self._label_en

    @label_en.setter
    def label_en(self, value: str):
        """ setter: label_en """
        if value:
            self._label_en = value.strip()
        else:
            self._label_en = ''

    @property
    def statement_property(self) -> str:
        """ get statement_property """
        return self._statement_property

    @statement_property.setter
    def statement_property(self, value: str):
        """ set statement_property"""
        if value:
            self._statement_property = value.strip()
        else:
            self._statement_property = ''

    @property
    def statement_value(self):
        """ get statement_value """
        return self._statement_value

    @statement_value.setter
    def statement_value(self, value: str):
        """ set statement_value """
        if value:
            self._statement_value = value.strip()
        else:
            self._statement_value = ''

    def write_to_wikibase(self):
        """ zapis deklaracji dla elementu w instancji wikibase
            także zapis aliasu dla elementu - zależnie od wartości
            self.statement_property
        """
        is_ok, p_id = find_name_qid(self.label_en, 'item')
        if not is_ok:
            print('ERROR:', f'brak elementu -> {self.label_en}')
            return

        # jeżeli to alias?
        if self.statement_property in ('Apl', 'Aen', 'Ade', 'Aru', 'Aes', 'Afr', 'Alt', 'Alv', 'Aet', 
                                       'Anl', 'Ait', 'Ala', 'Ahu', 'Apt', 'Auk', 'Acs',
                                       'Ask', 'Asl', 'Aro', 'Asv', 'Afi'):
            try:
                wd_item = wbi_core.ItemEngine(item_id=p_id)
                lang = self.statement_property[1:]
                aliasy = wd_item.get_aliases(lang=lang)
                skip_alias = False
                if self.statement_value in aliasy:
                    print(f"SKIP: element: '{p_id}' już posiada alias: '{self.statement_value}' dla języka: {lang}.")
                    skip_alias = True
                else:
                    wd_item.set_aliases(self.statement_value, lang=self.statement_property[-2:])
                    wd_item.write(login_instance, entity_type='item')
                    print(f'ALIAS ADDED, item {p_id}: {self.statement_property} -> {self.statement_value}')

                # aliasy dla elementów powinny od razu stawać się także deklaracjami właściwości
                # 'stated as', ale tylko jeżeli są dla arkusza globalne referencje
                if not skip_alias and self.additional_references:
                    is_ok, prop_id = find_name_qid('stated as', 'property')
                    if not is_ok:
                        print('ERROR:', 'brak właściwości -> stated as')
                        return

                    lang_id = self.statement_property[1:]
                    p_value = f'{lang_id}:"{self.statement_value}"'
                    
                    # kontrola czy istnieje deklaracja o takiej wartości
                    if has_statement(p_id, prop_id, value_to_check=p_value):
                        print(f"SKIP: element: '{p_id}' już posiada deklarację: '{prop_id}' o wartości: {p_value}.")
                    else:
                        # wartości deklaracji 'stated as' są dołączane do istniejących, nie zastępują poprzednich!
                        st_data = create_statement_data(prop_id, p_value, self.references,
                                                    None, add_ref_dict=self.additional_references,
                                                    if_exists='APPEND')
                        if st_data:
                            try:
                                data =[st_data]
                                wd_statement = wbi_core.ItemEngine(item_id=p_id, data=data, debug=False)
                                #wd_statement.init_data_load()
                                #wd_statement.update(data)
                                wd_statement.write(login_instance, entity_type='item')

                                print(f'STATEMENT ADDED, {p_id}: {prop_id} -> {p_value}')
                            except (MWApiError, KeyError, ValueError):
                                print(f'ERROR, {p_id}: {prop_id} -> {p_value}')
                        else:
                            print(f'INVALID DATA, {p_id}: {prop_id} -> {p_value}')

            except (MWApiError, KeyError, ValueError):
                print(f'ERROR: item {p_id} {self.statement_property} -> {self.statement_value}')

        # jeżeli to etykieta
        elif self.statement_property in ('Lde', 'Lru', 'Les', 'Lfr', 'Llt', 'Llv', 'Let',
                                         'Lnl', 'Lit', 'Lla', 'Lhu', 'Lpt', 'Luk', 'Lcs',
                                         'Lsk', 'Lsl', 'Lro', 'Lsv', 'Lfi'):
            try:
                wd_item = wbi_core.ItemEngine(item_id=p_id)
                wd_item.set_label(self.statement_value, lang=self.statement_property[-2:], if_exists='REPLACE')
                wd_item.write(login_instance, entity_type='item')
                print(f'LABEL ADDED/MODIFIED, item {p_id}: {self.statement_property} -> {self.statement_value}')
            
            except (MWApiError, KeyError, ValueError):
                print(f'ERROR: item {p_id} {self.statement_property} -> {self.statement_value}')

        # jeżeli to deklaracja?
        else:
            is_ok, prop_id = find_name_qid(self.statement_property, 'property')
            if not is_ok:
                print('ERROR:', f'w instancji wikibase brak właściwości -> {self.statement_property}')
                return

            if self.qualifiers:
                # zmiana nazwy kwalifikatora na jego Q
                tmp = {}
                for q_key, value in self.qualifiers.items():
                    is_ok, qualifier_id = find_name_qid(q_key, 'property')
                    if not is_ok:
                        print('ERROR:', f'w instancji wikibase brak właściwości -> {q_key}')
                        return

                    tmp[qualifier_id] = value

                self.qualifiers = tmp
    
            # tu obsługa specyficznych typów właściwości: item/property wartość
            # wprowadzana jako deklaracją powinna być symbolem P lub Q
            prop_type = get_property_type(prop_id)
            if prop_type == 'wikibase-item':
                is_ok, p_value = find_name_qid(self.statement_value, 'item')
                if not is_ok:
                    print('ERROR:', f'brak elementu -> {self.statement_value} będącego wartością -> {self.statement_property}')
                    return
            elif prop_type == 'wikibase-property':
                is_ok, p_value = find_name_qid(self.statement_value, 'property')
                if not is_ok:
                    print('ERROR:', f'brak właściwości -> {self.statement_value} będącej wartością -> {self.statement_property}')
                    return
            else:
                p_value = self.statement_value

            # tu podobna obsługa j.w. ale tym razem dla dla kwalifikatorów
            tmp = {}
            for key, value in self.qualifiers.items():
                qualifier_type = get_property_type(key)
                if qualifier_type == 'wikibase-item':
                    is_ok, q_value = find_name_qid(value, 'item')
                    if not is_ok:
                        print('ERROR:', f'brak elementu -> {value} będącego wartością kwalifikatora -> {key}')
                        return
                elif qualifier_type == 'wikibase-property':
                    is_ok, q_value = find_name_qid(value, 'property')
                    if not is_ok:
                        print('ERROR:', f'brak właściwości -> {value} będącej wartością kwalifikatora -> {key}')
                        return
                else:
                    q_value = value

                tmp[key] = q_value

            self.qualifiers = tmp
            
            # kontrola czy istnieje deklaracja o tej wartości
            if has_statement(p_id, prop_id, value_to_check=p_value):
                print(f"SKIP: element: '{p_id}' już posiada deklarację: '{prop_id}' o wartości: {p_value}.")
            else:
                st_data = create_statement_data(prop_id, p_value, self.references,
                                                self.qualifiers, add_ref_dict=self.additional_references,
                                                if_exists='APPEND')
                if st_data:
                    try:
                        data =[st_data]
                        wd_statement = wbi_core.ItemEngine(item_id=p_id, data=data, debug=False)
                        wd_statement.write(login_instance, entity_type='item')
                        print(f'STATEMENT ADDED, {p_id}: {prop_id} -> {self.statement_value}')
                    except (MWApiError, KeyError, ValueError):
                        print(f'ERROR, {p_id}: {prop_id} -> {self.statement_value}')
                else:
                    print(f'INVALID DATA, {p_id}: {prop_id} -> {self.statement_value}')


# --- funkcje ---

def add_property(p_dane: WDHProperty) -> tuple:
    """
    funkcja dodaje nową właściwość
    zwraca tuple: (True/False, ID/ERROR)
    """

    # test czy właściwość już nie istnieje
    search_property, search_id = element_search(p_dane.label_en, 'property', 'en')
    if search_property:
        print(f"Property: '{p_dane.label_en}' already exists: {search_id}, update mode.")
        wd_item = wbi_core.ItemEngine(item_id=search_id)
        mode = 'updated: '
    else:
        wd_item = wbi_core.ItemEngine(new_item=True)
        mode = 'added: '

    # etykiety i opisy
    wd_item.set_label(p_dane.label_en, lang='en')
    wd_item.set_description(p_dane.description_en, lang='en')
    if p_dane.label_pl:
        wd_item.set_label(p_dane.label_pl,lang='pl')
    if p_dane.description_pl:
        wd_item.set_description(p_dane.description_pl, lang='pl')

    # Wikidata ID i Wikidata URL
    wiki_dane = None
    if p_dane.wiki_id:
        if wikibase_prop.wiki_id == '' or wikibase_prop.wiki_url == '':
            wikibase_prop.get_wiki_properties()

        url = f"https://www.wikidata.org/wiki/Property:{p_dane.wiki_id}"
        references = [
            [
                wbi_datatype.Url(value=url, prop_nr=wikibase_prop.wiki_url, is_reference=True)
            ]
        ]
        wiki_dane = wbi_datatype.ExternalID(value=p_dane.wiki_id, prop_nr=wikibase_prop.wiki_id,
            references=references)

    # odwrotność właściwości
    inverse_dane = None
    if p_dane.inverse_property:
        if wikibase_prop.inverse == '':
            wikibase_prop.get_wiki_properties()
        search_inverse, inv_pid = element_search(p_dane.inverse_property, 'property', 'en')
        if search_inverse and wikibase_prop.inverse != '':
            inverse_dane = wbi_datatype.Property(value=inv_pid, prop_nr=wikibase_prop.inverse)

    # typy danych dla property: 'string', 'wikibase-item', 'wikibase-property',
    # 'monolingualtext', 'external-id', 'quantity', 'time', 'geo-shape', 'url',
    # 'globe-coordinate'
    options = {'property_datatype':p_dane.datatype}

    try:
        p_new_id = wd_item.write(login_instance, entity_type='property', **options)
        if search_property:
            p_new_id = search_id

        # deklaracje dla właściwości
        data = []
        if wiki_dane:
            data.append(wiki_dane)
        if inverse_dane:
            data.append(inverse_dane)

        if len(data) > 0:
            wd_statement = wbi_core.ItemEngine(item_id=p_new_id, data=data, debug=False)
            wd_statement.write(login_instance, entity_type='property')

        # jeżeli dodano właściwość inverse_property do dla docelowej właściwości należy
        # dodać odwrotność: nową właściwość jako jej inverse_property
        if inverse_dane:
            inv_statement = WDHStatementProperty(inv_pid, wikibase_prop.inverse, p_new_id)
            add_res, add_info = add_property_statement(inv_statement)
            if not add_res:
                print(f'{add_info}')

        add_result = (True, mode + p_new_id)

    except (MWApiError, KeyError):
        add_result = (False, 'ERROR')

    return add_result


def find_name_qid(name: str, elem_type: str) -> tuple:
    """Funkcja sprawdza czy przekazany argument jest identyfikatorem właściwości/elementu
       jeżeli nie to szuka w wikibase właściwości/elementu o etykiecie (ang) równej argumentowi
       i zwraca jej id
    """
    output = (True, name)               # zakładamy, że w name jest id (np. P47)
                                        # ale jeżeli nie, to szukamy w wikibase
    if elem_type == 'property':
        pattern = r'^P\d{1,9}$'
    elif elem_type == 'item':
        pattern = r'^Q\d{1,9}$'

    match = re.search(pattern, name)
    if not match:
        output = element_search(name, elem_type, 'en')
        if not output[0]:
            output =  (False, f'INVALID DATA, {elem_type}: {name}, {output[1]}')

    return output


def create_statement(prop: str, value: str, is_ref: bool = False, refs = None,
                                            is_qlf: bool = False, qlfs = None,
                                            if_exists:str = 'REPLACE') ->Union[
                                                       wbi_datatype.String,
                                                       wbi_datatype.Property,
                                                       wbi_datatype.ItemID,
                                                       wbi_datatype.ExternalID,
                                                       wbi_datatype.Url,
                                                       wbi_datatype.Quantity,
                                                       wbi_datatype.Time,
                                                       wbi_datatype.GeoShape,
                                                       wbi_datatype.GlobeCoordinate,
                                                       wbi_datatype.MonolingualText]:
    """
    Funkcja tworzy obiekt będący deklaracją lub referencją lub kwalifikatorem
    """
    statement = None

    res, property_nr = find_name_qid(prop, 'property')
    if res:
        property_type = get_property_type(property_nr)
        if property_type == 'string':
            statement = wbi_datatype.String(value=value, prop_nr=property_nr,
                                            is_reference=is_ref, references=refs,
                                            is_qualifier=is_qlf, qualifiers=qlfs, 
                                            if_exists=if_exists)
        elif property_type == 'wikibase-item':
            res, value_id = find_name_qid(value, 'item')
            if res:
                statement = wbi_datatype.ItemID(value=value_id, prop_nr=property_nr,
                                                is_reference=is_ref, references=refs,
                                                is_qualifier=is_qlf, qualifiers=qlfs, 
                                                if_exists=if_exists)
        elif property_type == 'wikibase-property':
            res, value_id = find_name_qid(value, 'property')
            if res:
                statement = wbi_datatype.Property(value=value_id, prop_nr=property_nr,
                                                  is_reference=is_ref, references=refs,
                                                  is_qualifier=is_qlf, qualifiers=qlfs, 
                                                  if_exists=if_exists)
        elif property_type == 'external-id':
            statement = wbi_datatype.ExternalID(value=value, prop_nr=property_nr,
                                                is_reference=is_ref, references=refs,
                                                is_qualifier=is_qlf, qualifiers=qlfs, 
                                                if_exists=if_exists)
        elif property_type == 'url':
            statement = wbi_datatype.Url(value=value, prop_nr=property_nr,
                                         is_reference=is_ref, references=refs,
                                         is_qualifier=is_qlf, qualifiers=qlfs, 
                                         if_exists=if_exists)
        elif property_type == 'monolingualtext':
            # zakładając że wartość monolingualtext jest zapisana w formie:
            # pl:"To jest tekst w języku polskim", a jeżeli brak przedrostka z kodem
            # języka to przyjmujemy 'en'
            if value[2] == ':':
                prop_lang = value[:2]
                value = value[4:-1] # bez cudzysłowów
            else:
                prop_lang = 'en'
            statement = wbi_datatype.MonolingualText(text=value, prop_nr=property_nr, language=prop_lang,
                                                     is_reference=is_ref, references=refs,
                                                     is_qualifier=is_qlf, qualifiers=qlfs, 
                                                     if_exists=if_exists)
        elif property_type == 'quantity':
            statement = wbi_datatype.Quantity(quantity=value, prop_nr=property_nr,
                                              is_reference=is_ref, references=refs,
                                              is_qualifier=is_qlf, qualifiers=qlfs, 
                                              if_exists=if_exists)
        elif property_type == 'time':
            # czas w formacie +1539-12-08T00:00:00Z/11, po slashu precyzja daty zgodnie
            # ze standardami wikibase 11 - dzień, 9 - rok
            tmp = value.split("/")
            if len(tmp) == 2:
                time_value = tmp[0]
                precision = int(tmp[1])
                statement = wbi_datatype.Time(time_value, prop_nr=property_nr,
                                              precision=precision, is_reference=is_ref,
                                              references=refs, is_qualifier=is_qlf,
                                              qualifiers=qlfs, if_exists=if_exists)
            else:
                print(f'ERROR: invalid value for time type: {value}.')
        elif property_type == 'geo-shape':
            # to chyba oczekuje nazwy pliku mapy w wikimedia commons, nam się nie przyda?
            statement = wbi_datatype.GeoShape(value, prop_nr=property_nr, is_reference=is_ref,
                                              references=refs, is_qualifier=is_qlf,
                                              qualifiers=qlfs, if_exists=if_exists)
        elif property_type == 'globe-coordinate':
            # oczekuje na zapis w formacie: 51.2,20.1 opcjonalnie jeszcze ,0.1
            # czyli latitude, longitude (jako liczby dziesiętne), oraz precyzja, domyślnie 0.1
            # domyślny glob = Earth, ale można zmienić na Marsa
            # https://www.wikidata.org/wiki/Help:Data_type/pl#Globe_coordinate

            tmp = value.split(",")
            try:
                latitude = float(tmp[0])
                longitude = float(tmp[1])
                if len(tmp) > 2:
                    precision = float(tmp[2])
                else:
                    precision = 0.1
            except ValueError:
                print(f'ERROR: invalid value for globe-coordinate type: {value}.')
            else:
                statement = wbi_datatype.GlobeCoordinate(latitude, longitude, precision,
                                                         prop_nr=property_nr, is_reference=is_ref,
                                                         references=refs, is_qualifier=is_qlf,
                                                         qualifiers=qlfs, if_exists=if_exists)

    return statement


def create_references(ref_dict: dict, additional_ref_dict: dict = None, 
                      if_exists: str = 'REPLACE') ->list:
    """ Funkcja tworzy referencje z przekazanego słownika referencji, opcjonalnie
        może zostać przekazany drugi słownik referencji np. globalnych wówczas
        utworzny zostanie drugi blok referencji
    """
    if ref_dict:
        statements = []
        for key, value in ref_dict.items():
            statement = create_statement(key, value, is_ref=True, refs=None, if_exists=if_exists)
            if statement:
                statements.append(statement)
        new_references = [ statements ]
    else:
        new_references = None

    if additional_ref_dict:
        statements = []
        for key, value in  additional_ref_dict.items():
            statement = create_statement(key, value, is_ref=True, refs=None, if_exists=if_exists)
            if statement:
                statements.append(statement)
        
        if new_references:
            new_references.append(statements)
        else:
            new_references = [ statements ]

    return new_references


def create_qualifiers(qlf_dict: dict, if_exists: str = 'REPLACE') ->list:
    """ Funkcja tworzy kwalifikatory
    """
    new_qualifiers = []
    for key, value in qlf_dict.items():
        statement = create_statement(key, value, is_qlf=True, qlfs=None, if_exists=if_exists)
        if statement:
            new_qualifiers.append(statement)

    if len(new_qualifiers) == 0:
        new_qualifiers = None

    return new_qualifiers


def create_statement_data(prop: str, value: str, reference_dict: dict,
                          qualifier_dict: dict, add_ref_dict: dict = None,
                          if_exists: str = 'REPLACE') -> Union[
                                                       wbi_datatype.String,
                                                       wbi_datatype.Property,
                                                       wbi_datatype.ItemID,
                                                       wbi_datatype.ExternalID,
                                                       wbi_datatype.Url,
                                                       wbi_datatype.Quantity,
                                                       wbi_datatype.Time,
                                                       wbi_datatype.GeoShape,
                                                       wbi_datatype.GlobeCoordinate,
                                                       wbi_datatype.MonolingualText]:
    """
    Funkcja tworzy dane deklaracji z opcjonalnymy referencjami
    """
    # referencje i kwalifikatory z domyślną wartością if_exists = 'REPLACE'
    references = None
    if reference_dict or add_ref_dict:
        references = create_references(reference_dict, add_ref_dict)

    qualifiers = None
    if qualifier_dict:
        qualifiers = create_qualifiers(qualifier_dict)

    output_data = create_statement(prop, value, is_ref=False, refs=references,
                                   is_qlf=False, qlfs=qualifiers, if_exists=if_exists)

    return output_data


def add_property_statement(s_item: WDHStatementProperty) -> tuple:
    """
    Funkcja dodaje deklaracje (statement) do właściwości
    Parametry:
        s_item - obiekt z deklaracją
    """
    is_ok, p_id = find_name_qid(s_item.label_en, 'property')
    if not is_ok:
        return (False, p_id)

    is_ok, prop_id = find_name_qid(s_item.statement_property, 'property')
    if not is_ok:
        return (False, prop_id)

    # tu obsługa specyficznych typów właściwości: item/property wartość
    # wprowadzana jako deklaracją powinna być symbolem P lub Q
    prop_type = get_property_type(prop_id)
    if prop_type == 'wikibase-item':
        is_ok, value = find_name_qid(s_item.statement_value, 'item')
        if not is_ok:
            return (False, value)
    elif prop_type == 'wikibase-property':
        is_ok, value = find_name_qid(s_item.statement_value, 'property')
        if not is_ok:
            return (False, value)
    else:
        value = s_item.statement_value

    # kontrola czy istnieje deklaracja o takiej wartości
    if has_statement(p_id, prop_id, value_to_check=value):
        return (False, f"SKIP: property: '{p_id}' already has a statement: '{prop_id} with value: {value}'.")

    st_data = create_statement_data(s_item.statement_property, value,
                                    s_item.references,
                                    qualifier_dict=None,
                                    add_ref_dict=s_item.additional_references, if_exists='APPEND')
    if st_data:
        try:
            data =[st_data]
            wd_statement = wbi_core.ItemEngine(item_id=p_id, data=data, debug=False)
            wd_statement.write(login_instance, entity_type='property')
            add_result = (True, f'STATEMENT ADDED, {p_id}: {prop_id} -> {s_item.statement_value}')
        except (MWApiError, KeyError, ValueError):
            add_result = (False, f'ERROR, {p_id}: {prop_id} -> {s_item.statement_value}')
    else:
        add_result = (False, f'INVALID DATA, {p_id}: {prop_id} -> {s_item.statement_value}')

    return add_result


def get_property_type(p_id: str) -> str:
    """ Funkcja zwraca typ właściwości na podstawie jej identyfikatora
    """
    params = {'action': 'wbgetentities', 'ids': p_id,
              'props': 'datatype'}

    search_results = mediawiki_api_call_helper(data=params, login=None, mediawiki_api_url=None,
                                               user_agent=None, allow_anonymous=True)
    data_type = None
    if search_results:
        data_type = search_results['entities'][p_id]['datatype']

    return data_type


def has_statement(pid_to_check: str, claim_to_check: str, value_to_check: str=''):
    """
    Funkcja weryfikuje czy właściwość (property) lub element (item) ma już
    taką deklarację (statement), opcjonalnie - z podaną wartością
    """
    has_claim = False
    wb_prop = wbi_core.ItemEngine(item_id=pid_to_check)
    data_prop = wb_prop.get_json_representation()
    claims = data_prop['claims']
    if claim_to_check in claims:
        if not value_to_check:
            has_claim = True
        else:
            lista = claims[claim_to_check]
            for item in lista:
                value_json = item['mainsnak']['datavalue']['value']

                if 'type' in item['mainsnak']['datavalue'] and item['mainsnak']['datavalue']['type'] == 'string':
                    value = value_json
                elif 'text' in value_json and 'language' in value_json:
                    # jeżeli nietypowy cudzysłów w wartości z arkusza xlsx 
                    if '”' in value_to_check:
                        value_to_check = value_to_check.replace('”', '"')
                    value = f"{value_json['language']}:\"{value_json['text']}\""
                elif 'entity-type' in value_json:
                    value = value_json['id']
                elif 'latitude' in value_json:
                    value = f"{value_json['latitude']},{value_json['longitude']}"
                elif 'time' in value_json:
                    value = f"{value_json['time']}/{value_json['precision']}"
                else:
                    value = '???'
                    print(item, ' - ', value_to_check)

                if value == value_to_check:
                    has_claim = True
                    break

    return has_claim


if __name__ == "__main__":
    # login i hasło ze zmiennych środowiskowych
    env_path = Path('.') / '.env'
    load_dotenv(dotenv_path=env_path)

    BOT_LOGIN = os.environ.get('WIKIDARIAH_USER')
    BOT_PASSWORD = os.environ.get('WIKIDARIAH_PWD')

    login_instance = wbi_login.Login(user=BOT_LOGIN, pwd=BOT_PASSWORD)

    # podstawowe właściwości Wikibase
    wikibase_prop = BasicProp()

    # ustalenie nr podstawowych property (jeżeli są, jeżeli będą dodawane podczas
    # pracy skryptu, wartości zostaną podczytane pred pierwszym użyciem)
    wikibase_prop.get_wiki_properties()

    # dane z arkusza XLSX, wg ścieżki przekazanej argumentem z linii komend
    # jeżeli nie przekazano, skrypt szuka pliku 'data/arkusz_import.xlsx'
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    else:
        filename = Path('.') / 'data/arkusz_import.xlsx'

    plik_xlsx = WDHSpreadsheet(filename)
    plik_xlsx.open()

    # globalne referencje
    plik_xlsx.get_global()

    # właściwośći
    dane = plik_xlsx.get_property_list()
    for wb_property in dane:
        result, info = add_property(wb_property)
        if result:
            print(f'Property {info}')

    # dodatkowe deklaracje dla właściwości
    dane = plik_xlsx.get_statement_list()
    for stm in dane:
        result, info = add_property_statement(stm)
        print(f'{info}')

    # elementy 'strukturalne' ('definicyjne')
    dane = plik_xlsx.get_item_list()
    for wb_item in dane:
        wb_item.write_to_wikibase()

    # dodatkowe deklaracje dla elementów strukturalnych/definicyjnych
    dane = plik_xlsx.get_item_statement_list()
    for stm in dane:
        stm.write_to_wikibase()
